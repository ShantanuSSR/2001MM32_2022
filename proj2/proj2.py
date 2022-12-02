from datetime import datetime

start_time = datetime.now()
import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.styles import Color, PatternFill, Font, Border, Side
import base64
import io
import tkinter as tk
from tkinter import filedialog
import glob
import os
from zipfile import ZipFile
from io import BytesIO

from platform import python_version

ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print(
        "Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

order_label = [1, -1, 2, -2, 3, -3, 4, -4]


MOD = 0
file_uploaded = None


def proj2_Front_End_Interface():
    st.set_page_config(page_title="Project 2", layout="wide")
    st.title("Project 2")
    hide_default_format = """
       <style>
       #MainMenu {visibility: hidden; }
       footer {visibility: hidden;}
       </style>
       """
    st.markdown(hide_default_format, unsafe_allow_html=True)
    
    st.header("Welcome to the FrontEnd Interface")
    folderPath = ""

    option = st.radio("Select conversion type: ", ('Single file conversion', 'Bulk file conversion'))

    if option == "Single file conversion":
        global uploadedFile
        uploadedFile = st.file_uploader("Upload File", type=['xlsx'], accept_multiple_files=False, key="fileUploader")

        if "folderPath" in st.session_state:
            del st.session_state["folderPath"]

    if option == "Bulk file conversion":
        uploadedFile = None

        # Set up tkinter
        root = tk.Tk()
        root.withdraw()

        # Making Select Folder for bulk conversion dialog appear on top of other windows
        root.wm_attributes('-topmost', 1)

        # Select Folder for bulk conversion button
        st.write('Please select a folder:')
        clicked = st.button('Select Folder for bulk conversion')
        if clicked:
            folderPath = filedialog.askdirectory(master=root)
            st.session_state["folderPath"] = folderPath

    if "folderPath" in st.session_state:
        folderPath = st.session_state["folderPath"]
        dirname = st.text_input('Selected folder:', folderPath)

    global MOD
    MOD = st.number_input('Enter MOD value: ', min_value=1, value=5000, step=1)

    convert, download = st.columns(2)

    with convert:
        conv = st.button("Compute")

        if conv:
            if option == "Single file conversion":
                if not uploadedFile:
                    st.warning("Please upload a xlsx file")
                else:
                    fileName = uploadedFile.name.split(".xlsx")[0]
                    outputFileName = startConversion(fileName)

                    with download:
                        with open(outputFileName, 'rb') as my_file:
                            st.download_button(label='Download File', data=my_file, file_name=outputFileName,
                                               mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            elif option == "Bulk file conversion":
                if "folderPath" in st.session_state:
                    folderPath = st.session_state["folderPath"]

                folder = folderPath.split("/")[-1]

                excel_files = glob.glob(os.path.join(folderPath, "*.xlsx"))

                if len(folder) == 0:
                    st.warning("Please select a folder")
                    return

                if len(excel_files) == 0:
                    st.warning("No excel sheet found!!")
                    return

                outputFolderName = getOutputFileName(folder) + ".zip"

                zipObj = ZipFile(outputFolderName, "w")

                for i, file in enumerate(excel_files):
                    uploadedFile = file
                    fileName = file.split(".xlsx")[0]
                    fileName = fileName.split("\\")[-1]

                    outputFileName = startConversion(fileName)

                    zipObj.write(outputFileName)

                zipObj.close()

                with download:
                    with open(outputFolderName, 'rb') as my_file:
                        st.download_button(label="Download result", data=my_file, file_name=outputFolderName)


def startConversion(fileName):
    df = pd.read_excel(uploadedFile)


    output_file = getOutputFileName(fileName) + ".xlsx"

    outputFile = openpyxl.Workbook()
    currentSheet = outputFile.active


    overall_count = 0

    colmn = 1

    # storing variables for the sum
    u_sum = 0
    v_sum = 0
    w_sum = 0

    for key, value in df.items():
        value = value.tolist()
        overall_count = len(value)

        # i -> 2nd row
        currentSheet.cell(row=2, column=colmn).value = key

        for r, val in enumerate(value):
            if colmn == 2:
                u_sum += val
            elif colmn == 3:
                v_sum += val
            elif colmn == 4:
                w_sum += val

            currentSheet.cell(row=r + 3, column=colmn).value = val

        colmn += 1

    # computing the average

    u_avg = round(u_sum / overall_count, 3)
    v_avg = round(v_sum / overall_count, 3)
    w_avg = round(w_sum / overall_count, 3)


    # defining average values

    currentSheet.cell(row=3, column=5).value = u_avg
    currentSheet.cell(row=3, column=6).value = v_avg
    currentSheet.cell(row=3, column=7).value = w_avg


    # Processing input
    SetOctantProcessedData(u_avg, v_avg, w_avg, currentSheet)

    currentSheet.cell(row=1, column=14).value = "Overall Octant Count"
    currentSheet.cell(row=1, column=24).value = "Rank #1 Should be highlighted Yellow"
    currentSheet.cell(row=1, column=35).value = "Overall Transition Count"
    currentSheet.cell(row=1, column=45).value = "Longest Subsequence Length"
    currentSheet.cell(row=1, column=49).value = "Longest Subsequence Length with Range"

    currentSheet.cell(row=2, column=36).value = "To"

    row1 = ["T", "U", "V", "W", "U Avg", "V Avg", "W Avg", "U'=U - U avg", "V'=V - V avg", "W'=W - W avg", "Octant"]
    for i, head in enumerate(row1):
        currentSheet.cell(row=2, column=i + 1).value = head

    OctantOverallRankFormat(currentSheet, MOD, overall_count)

    FormatCountModWise(currentSheet, MOD, overall_count)

    TransitionOverallCountFormat(currentSheet, overall_count)

    # # Function to add a mod-wise transition count
    TransitionCountModWiseFormat(currentSheet, MOD, overall_count)

    LongestSubsequenceChecker(currentSheet, overall_count)

    outputFile.save(output_file)

    data = currentSheet.values
    columns = next(data)[0:]

    df = pd.DataFrame(data, columns=columns)

    return output_file


def getOutputFileName(inputFile):
    output_file = inputFile + "mod" + str(MOD) + "_"
    now = datetime.now()
    dt_string = now.strftime("%Y-%m-%d-%H-%M-%S")
    output_file += dt_string

    return output_file


def InitiateCount(count):
    for x in order_label:
        count[x] = 0


# A method to initialize a dictionary with 0 for all "order label" values other than "left"
def InitiateCountExcept(count, left):
    for item in order_label:
        if (item != left):
            count[item] = 0


def LongestSubsequenceChecker(currentSheet, overall_count):
    # Dictionary to store the number of consecutive sequences
    count = {}

    # Dictionary to store longest count
    longest = {}

    # all labels in the dictionary are initialised to 0.
    InitiateCount(count)
    InitiateCount(longest)

    # variable to check the previous value
    last = -10

    # Iterating over the complete excel sheet
    for i in range(overall_count):
        presnRow = i + 3

        curr = int(currentSheet.cell(column=11, row=presnRow).value)

        # Comparing current and last value
        if (curr == last):
            count[curr] += 1
            longest[curr] = max(longest[curr], count[curr])
            InitiateCountExcept(count, curr)
        else:
            count[curr] = 1
            longest[curr] = max(longest[curr], count[curr])
            InitiateCountExcept(count, curr)


        # Updating "last" variable
        last = curr

    # Longest subsequence frequency counting method
    LongestSubsequenceFrequencyChecker(longest, currentSheet, overall_count)


def LongestSubsequenceFrequencyChecker(longest, currentSheet, overall_count):
    #Dictionary to store the number of consecutive sequence count
    count = {}

    # Dictinary to store the number of frequency count
    frequency = {}

    # Dictionary to store time range
    timerange_dict = {}

    for label in order_label:
        timerange_dict[label] = []

    # the dictionary are initialized to 0.
    InitiateCount(count)
    InitiateCount(frequency)

    # variable to check the last value
    last = -10

    # full Excel sheet iterations
    for i in range(overall_count):
        presnRow = i + 3
        try:
            curr = int(currentSheet.cell(column=11, row=presnRow).value)

            # contrasting the present and last values
            if (curr == last):
                count[curr] += 1
            else:
                count[curr] = 1
                InitiateCountExcept(count, curr)

            # Upading frequency
            if (count[curr] == longest[curr]):
                frequency[curr] += 1

                # counting the beginning and finish of the longest sequence
                end = float(currentSheet.cell(row=presnRow, column=1).value)
                start = 100 * end - longest[curr] + 1
                start /= 100

                # Adding a time interval to a map
                timerange_dict[curr].append([start, end])

                # Resetting the dictionary count
                InitiateCount(count)
            else:
                InitiateCountExcept(count, curr)
        except FileNotFoundError:
            print("File not found!!")
            exit()
        except ValueError:
            print("File content is invalid!!")
            exit()

        # Updating 'last' variable
        last = curr

    # Setting frequency table into sheet
    FrequencyChecker(longest, frequency, currentSheet)

    # Setting the time range for the longest sequence
    FormatLongestSubsequenceTime(longest, frequency, timerange_dict, currentSheet)


# Method to set frequency count to sheet
def FrequencyChecker(longest, frequency, currentSheet):
    # updating sheet and iterating "order label"
    # for i in range(9):
    i = 0
    j = 0
    while i < 9:
        # for j in range(3):
        while j < 3:
            black = "00000000"
            double = Side(border_style="thin", color=black)
            black_border = Border(top=double, left=double, right=double, bottom=double)
            currentSheet.cell(row=3 + i, column=45 + j).border = black_border
            j += 1
        i += 1

    currentSheet.cell(row=3, column=45).value = "Octant ##"
    currentSheet.cell(row=3, column=46).value = "Longest Subsquence Length"
    currentSheet.cell(row=3, column=47).value = "Count"

    for i, label in enumerate(order_label):
        presnRow = i + 3

        currentSheet.cell(row=presnRow + 1, column=45).value = label
        currentSheet.cell(column=46, row=presnRow + 1).value = longest[label]
        currentSheet.cell(column=47, row=presnRow + 1).value = frequency[label]



# Method to set time range for longest subsequence
def FormatLongestSubsequenceTime(longest, frequency, timerange_dict, currentSheet):
    # Naming columns number
    colmn_length = 50
    freq_col = 51

    # Initial row, just after the head row
    row = 4

    currentSheet.cell(row=3, column=49).value = "Octant ###"
    currentSheet.cell(row=3, column=50).value = "Longest Subsquence Length"
    currentSheet.cell(row=3, column=51).value = "Count"

    # Iterating all octants
    for octant in order_label:

        # Setting octant's longest subsequence and frequency data
        currentSheet.cell(column=49, row=row).value = octant
        currentSheet.cell(column=colmn_length, row=row).value = longest[octant]
        currentSheet.cell(column=freq_col, row=row).value = frequency[octant]


        row += 1


        # Setting default labels
        currentSheet.cell(column=49, row=row).value = "Time"
        currentSheet.cell(column=colmn_length, row=row).value = "From"
        currentSheet.cell(column=freq_col, row=row).value = "To"


        row += 1

        # time range values for each octant are iterated.
        for timeData in timerange_dict[octant]:
            try:
                # Setting time interval value
                currentSheet.cell(row=row, column=colmn_length).value = timeData[0]
                currentSheet.cell(row=row, column=freq_col).value = timeData[1]
            except FileNotFoundError:
                print("File not found!!")
                exit()
            row += 1

    black = "00000000"
    double = Side(border_style="thin", color=black)
    black_border = Border(top=double, left=double, right=double, bottom=double)

    for i in range(3, row):
        for j in range(49, 52):
            currentSheet.cell(row=i, column=j).border = black_border


def TransitionCountModWiseFormat(currentSheet, mod, overall_count):
    # partition counting in respect to mod
    try:
        overall_partition = overall_count // mod
    except ZeroDivisionError:
        print("Mod can't have 0 value")
        exit()

    # Checking mod value range
    if (mod < 0):
        raise Exception("Mod value should be in range of 1-30000")

    if (overall_count % mod != 0):
        overall_partition += 1

    # row starts are initialized for data filling.
    row_begin = 16

    # Iterating all partitions
    for i in range(overall_partition):
        # Initialising start and end values
        start = i * mod
        end = min((i + 1) * mod - 1, overall_count - 1)

        # Setting start-end values
        try:
            currentSheet.cell(column=35, row=row_begin - 1 + 13 * i).value = "Mod Transition Count"
            currentSheet.cell(column=35, row=row_begin + 13 * i).value = str(start) + "-" + str(end)
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        # Initialising empty dictionary
        transition_count = {}
        for a in range(1, 5):
            for b in range(1, 5):
                transition_count[str(a) + str(b)] = 0
                transition_count[str(a) + str(-b)] = 0
                transition_count[str(-a) + str(b)] = 0
                transition_count[str(-a) + str(-b)] = 0

        # Counting transition for range [start, end)
        for a in range(start, end + 1):
            try:
                curr = currentSheet.cell(column=11, row=a + 3).value
                next = currentSheet.cell(column=11, row=a + 4).value
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

            # counting up for values within a range
            if (next != None):
                transition_count[str(curr) + str(next)] += 1

        # Setting transition counts
        setTransitionCount(row_begin + 13 * i, transition_count, currentSheet)


def TransitionOverallCountFormat(currentSheet, overall_count):
    # Setting value

    # Initialising empty dictionary
    transition_count = {}
    for i in range(1, 5):
        for j in range(1, 5):
            transition_count[str(i) + str(j)] = 0
            transition_count[str(i) + str(-j)] = 0
            transition_count[str(-i) + str(j)] = 0
            transition_count[str(-i) + str(-j)] = 0

    # using iteration to fill the dictionary with octant values
    start = 0

    # For the string to int conversion, try and except block.
    try:
        last = int(currentSheet["K3"].value)
    except ValueError:
        print("Sheet input can't be converted to int")
        exit()
    except TypeError:
        print("Sheet doesn't contain integer octant")
        exit()

    while (start < overall_count - 1):
        # try and except block for string to int conversion
        try:
            curr = int(currentSheet.cell(row=start + 4, column=11).value)
            transition_count[str(last) + str(curr)] += 1
            last = curr
        except ValueError:
            print("Sheet input can't be converted to int")
            exit()
        except TypeError:
            print("Sheet doesn't contain integer octant")
            exit()

        start += 1

    # Setting transitions counted into sheet
    setTransitionCount(2, transition_count, currentSheet)


# to set transition count defining a setTransitionCount function
def setTransitionCount(row, transition_count, currentSheet):
    # Setting hard coded inputs

    currentSheet.cell(row=row, column=36).value = "To"
    currentSheet.cell(row=row + 1, column=35).value = "Octant #"
    currentSheet.cell(row=row + 2, column=34).value = "From"

    black = "00000000"
    double = Side(border_style="thin", color=black)
    black_border = Border(top=double, left=double, right=double, bottom=double)

    for i in range(35, 44):
        for j in range(row + 1, row + 1 + 9):
            currentSheet.cell(row=j, column=i).border = black_border




    # Setting Labels
    for i, label in enumerate(order_label):
        try:
            currentSheet.cell(row=row + 1, column=i + 36).value = label
            currentSheet.cell(row=row + i + 2, column=35).value = label
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

    # Setting data
    for i, l1 in enumerate(order_label):
        max_i = -1

        for j, l2 in enumerate(order_label):
            val = transition_count[str(l1) + str(l2)]
            max_i = max(max_i, val)

        yellow = "00FFFF00"
        YellowBg = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')

        for j, l2 in enumerate(order_label):
            try:
                currentSheet.cell(row=row + i + 2, column=36 + j).value = transition_count[str(l1) + str(l2)]
                if transition_count[str(l1) + str(l2)] == max_i:
                    max_i = -1
                    currentSheet.cell(row=row + i + 2, column=36 + j).fill = YellowBg
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()


def FormatCountModWise(currentSheet, mod, overall_count):
    # Initialising empty dictionary
    count = {-1: 0, 1: 0, -2: 0, 2: 0, -3: 0, 3: 0, -4: 0, 4: 0}

    # Variable to store last row
    finalRow = -1

    # to set count dictionary iterating theloop
    start = 0
    while (start < overall_count):
        try:
            count[int(currentSheet.cell(row=start + 3, column=11).value)] += 1
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        start += 1

        try:
            if (start % mod == 0):
                # Setting row data
                try:
                    row = 4 + start // mod
                    finalRow = row
                    currentSheet.cell(row=row, column=14).value = str(start - mod) + "-" + str(
                        min(overall_count, start - 1))

                    for i, label in enumerate(order_label):
                        currentSheet.cell(row=row, column=15 + i).value = count[label]

                    setRankCount(row, count, currentSheet)
                except FileNotFoundError:
                    print("Output file not found!!")
                    exit()
                except ValueError:
                    print("Row or column values must be at least 1 ")
                    exit()

                # Reset count values
                count = {-1: 0, 1: 0, -2: 0, 2: 0, -3: 0, 3: 0, -4: 0, 4: 0}
        except ZeroDivisionError:
            print("Mod cannot have a value of 0.")
            exit()

    try:
        if (start % mod != 0):
            # Setting row data
            try:
                row = 5 + start // mod
                finalRow = row
                currentSheet.cell(row=row, column=14).value = str(start - mod) + "-" + str(min(overall_count, start - 1))
                for i, label in enumerate(order_label):
                    currentSheet.cell(row=row, column=15 + i).value = count[label]

                setRankCount(row, count, currentSheet)
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

    except ZeroDivisionError:
        print("Mod can't have 0 value")
        exit()

    if (finalRow != -1):
        RankMapOverallCount(finalRow, currentSheet)


def RankMapOverallCount(finalRow, currentSheet):
    count = {-1: 0, 1: 0, -2: 0, 2: 0, -3: 0, 3: 0, -4: 0, 4: 0}

    row = 4
    while currentSheet.cell(row=row, column=29).value is not None:
        oct = int(currentSheet.cell(row=row, column=31).value)
        count[oct] += 1
        row += 1

    black = "00000000"
    double = Side(border_style="thin", color=black)
    black_border = Border(top=double, left=double, right=double, bottom=double)

    for i in range(9):
        for j in range(3):
            row = finalRow + 2 + i
            colmn = 29 + j
            currentSheet.cell(row=row, column=colmn).border = black_border

    currentSheet.cell(column=29, row=finalRow + 2).value = "Octant ID"
    currentSheet.cell(column=30, row=finalRow + 2).value = "Octant Name "
    currentSheet.cell(column=31, row=finalRow + 2).value = "Count of Rank 1 Mod Values"

    octant_name_id_mapping = {1: "Internal outward interaction", 2: "External Ejection", 3: "External inward interaction",
                              4: "Internal sweep", -1: "External outward interaction", -2: "Internal Ejection",
                              -3: "Internal inward interaction", -4: "External sweep"}
    for j, oct in enumerate(order_label):
        currentSheet.cell(column=29, row=finalRow + 3 + j).value = oct
        currentSheet.cell(column=30, row=finalRow + 3 + j).value = octant_name_id_mapping[oct]
        currentSheet.cell(column=31, row=finalRow + 3 + j).value = count[oct]


def OctantOverallRankFormat(currentSheet, mod, overall_count):
    row1 = ["Octant ID", 1, -1, 2, -2, 3, -3, +4, -4, "Rank Octant 1", "Rank Octant -1", "Rank Octant 2",
                "Rank Octant -2", "Rank Octant 3", "Rank Octant -3", "Rank Octant 4", "Rank Octant -4",
                "Rank1 Octant ID", "Rank1 Octant Name"]

    overall_rows = overall_count // mod + 1 + 1  # head + overall
    if overall_count % mod != 0:
        overall_rows += 1

    black = "00000000"
    double = Side(border_style="thin", color=black)
    black_border = Border(top=double, left=double, right=double, bottom=double)

    for i, head in enumerate(row1):
        for j in range(overall_rows):
            currentSheet.cell(row=3 + j, column=14 + i).border = black_border

    for i, head in enumerate(row1):
        currentSheet.cell(row=3, column=i + 14).value = head

    currentSheet.cell(row=4, column=13).value = "Mod " + str(mod)

    setOverallCount(overall_count, currentSheet)


def setOverallCount(overall_count, currentSheet):
    # Initialising count dictionary
    count = {-1: 0, 1: 0, -2: 0, 2: 0, -3: 0, 3: 0, -4: 0, 4: 0}

    # count dictionary data incrementation
    try:
        for i in range(3, overall_count + 3):
            count[int(currentSheet.cell(column=11, row=i).value)] = count[int(currentSheet.cell(column=11,
                                                                                                row=i).value)] + 1
    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Row/Column should be at least 1 else the sheet input cannot be converted to an integer")
        exit()
    except TypeError:
        print("Sheet doesn't contact valid octant value!!")
        exit()

    # data into a sheet
    for i, label in enumerate(order_label):
        try:
            currentSheet.cell(row=4, column=i + 15).value = count[label]
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Values in a row or column must be at least 1.")
            exit()

    setRankCount(4, count, currentSheet)


def setRankCount(row, mapCount, currentSheet):
    # Copying the count list to sort
    sortedCount = []
    count = []
    for label in order_label:
        count.append(mapCount[label])

    for ct in count:
        sortedCount.append(ct)

    sortedCount.sort(reverse=True)

    rank = []

    for i, el in enumerate(count):
        for j, ell in enumerate(sortedCount):
            if (ell == el):
                rank.append(j + 1)
                sortedCount[j] = -1
                break

    rank1Oct = -10

    yellow = "00FFFF00"
    YellowBg = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')

    for j in range(0, 8):
        currentSheet.cell(row=row, column=23 + j).value = rank[j]
        if (rank[j] == 1):
            rank1Oct = order_label[j]
            currentSheet.cell(row=row, column=23 + j).fill = YellowBg

    octant_name_id_mapping = {1: "Internal outward interaction", -1: "External outward interaction",
                              2: "External Ejection", -2: "Internal Ejection", 3: "External inward interaction",
                              -3: "Internal inward interaction", 4: "Internal sweep", -4: "External sweep"}
    currentSheet.cell(row=row, column=31).value = rank1Oct
    currentSheet.cell(row=row, column=32).value = octant_name_id_mapping[rank1Oct]


def SetOctantProcessedData(u_avg, v_avg, w_avg, currentSheet):
    start = 3
    time = currentSheet.cell(start, 1).value

    # Iterating throught sheet
    while (time != None):
        # Calculating processed data

        u1 = currentSheet.cell(start, 2).value - u_avg
        v1 = currentSheet.cell(start, 3).value - v_avg
        w1 = currentSheet.cell(start, 4).value - w_avg

        u1 = round(u1, 3)
        v1 = round(v1, 3)
        w1 = round(w1, 3)

        oct = getOctant(u1, v1, w1)

        # Setting processed data
        try:
            currentSheet.cell(row=start, column=8).value = u1
            currentSheet.cell(row=start, column=9).value = v1
            currentSheet.cell(row=start, column=10).value = w1
            currentSheet.cell(row=start, column=11).value = oct
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("The minimum value for a row or column must be 1.")
            exit()

        start +=  1
        try:
            time = currentSheet.cell(start, 1).value
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("The minimum value for a row or column must be 1.")
            exit()


# if-else method to retrieve the octant type
def getOctant(x, y, z):
    if (x >= 0 and y >= 0):
        if (z >= 0):
            return 1
        else:
            return -1

    if (x < 0 and y >= 0):
        if (z >= 0):
            return 2
        else:
            return -2

    if (x < 0 and y < 0):
        if (z >= 0):
            return 3
        else:
            return -3

    if (x >= 0 and y < 0):
        if (z >= 0):
            return 4
        else:
            return -4



proj2_Front_End_Interface()

# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
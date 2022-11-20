from datetime import datetime
start_time = datetime.now()
import os
os.system("cls")

import openpyxl
import glob

from openpyxl.styles import Color, PatternFill, Font, Border, Side

octant_sign = [1,-1,2,-2,3,-3,4,-4]
octant_name_id_mapping = {1:"Internal outward interaction", -1:"External outward interaction", 2:"External Ejection", -2:"Internal Ejection", 3:"External inward interaction", -3:"Internal inward interaction", 4:"Internal sweep", -4:"External sweep"}
yellow = "00FFFF00"
yellow_bg = PatternFill(start_color=yellow, end_color= yellow, fill_type='solid')
black = "00000000"
double = Side(border_style="thin", color=black)
black_border = Border(top=double, left=double, right=double, bottom=double)

#Code starts here
def reset_count(count):
    for item in octant_sign:
        count[item] = 0

# Method to initialise dictionary with 0 for "octant_sign" except 'left'
def reset_count_except(count, left):
    for item in octant_sign:
        if(item!=left):
            count[item] = 0

def set_frequency(longest, frequency, outputSheet):
    # Iterating "octant_sign" and updating sheet
    for i in range(9):
        for j in range(3):
            outputSheet.cell(row = 3+i, column = 45+j).border = black_border

    outputSheet.cell(row=3, column=45).value= "Octant ##"
    outputSheet.cell(row=3, column=46).value= "Longest Subsquence Length"
    outputSheet.cell(row=3, column=47).value= "Count"

    for i, label in enumerate(octant_sign):
        currRow = i+3
        outputSheet.cell(row=currRow+1, column=45).value = label	
        outputSheet.cell(column=46, row=currRow+1).value = longest[label]
        outputSheet.cell(column=47, row=currRow+1).value = frequency[label]

# Function to set time range for longest subsequence
def longest_subsequence_time(longest, frequency, timeRange, outputSheet):
    # Naming columns number
    lengthCol = 50
    freqCol = 51
    
    # Initial row, just after the header row
    row = 4

    outputSheet.cell(row=3, column = 49).value = "Octant ###"
    outputSheet.cell(row=3, column = 50).value = "Longest Subsquence Length"
    outputSheet.cell(row=3, column = 51).value = "Count"

    # Iterating all octants 
    for octant in octant_sign:
        try:
            # Setting octant's longest subsequence and frequency data
            outputSheet.cell(column=49, row=row).value = octant
            outputSheet.cell(column=lengthCol, row=row).value = longest[octant]
            outputSheet.cell(column=freqCol, row=row).value = frequency[octant]
        except FileNotFoundError:
            print("File not found!!")
            exit()

        row+=1

        try:
            # Setting default labels
            outputSheet.cell(column=49, row=row).value = "Time"
            outputSheet.cell(column=lengthCol, row=row).value = "From"
            outputSheet.cell(column=freqCol, row=row).value = "To"
        except FileNotFoundError:
            print("File not found!!")
            exit()

        row+=1

        # Iterating time range values for each octants
        for timeData in timeRange[octant]:
            try:
                # Setting time interval value
                outputSheet.cell(row=row, column=lengthCol).value = timeData[0]
                outputSheet.cell(row=row, column=freqCol).value = timeData[1]
            except FileNotFoundError:
                print("File not found!!")
                exit()
            row += 1

    for i in range(3, row):
        for j in range(49, 52):
            outputSheet.cell(row=i, column = j).border = black_border

def count_longest_subsequence_freq_func(longest, outputSheet, total_count):
    # Dictionary to store consecutive sequence count
    count = {}

    # Dictionary to store frequency count
    frequency = {}

    # Dictionary to store time range
    timeRange = {}

    for label in octant_sign:
        timeRange[label] = []

    # Initialing dictionary to 0 for all labels
    reset_count(count)
    reset_count(frequency)

    # Variable to check last value
    last = -10

    # Iterating complete excel sheet
    for i in range(0, total_count):
        currRow = i+3
        try:
            curr = int(outputSheet.cell(column=11, row=currRow).value)
            
            # Comparing current and last value
            if(curr==last):
                count[curr]+=1
            else:
                count[curr]=1        
                reset_count_except(count, curr)

            # Updating frequency
            if(count[curr]==longest[curr]):
                frequency[curr]+=1

                # Counting starting and ending time of longest subsequence
                end = float(outputSheet.cell(row=currRow, column=1).value)
                start = 100*end - longest[curr]+1
                start/=100

                # Inserting time interval into map
                timeRange[curr].append([start, end])

                # Resetting count dictionary
                reset_count(count)
            else:
                reset_count_except(count, curr)
        except FileNotFoundError:
            print("File not found!!")
            exit()
        except ValueError:
            print("File content is invalid!!")
            exit()

        # Updating 'last' variable
        last = curr

    # Setting frequency table into sheet
    set_frequency(longest, frequency, outputSheet)

    # Setting time range for longest subsequence
    longest_subsequence_time(longest, frequency, timeRange, outputSheet)

# Function to set frequency count to sheet
			
def find_longest_subsequence(outputSheet, total_count):
	# Dictionary to store consecutive sequence count
    count = {}

    # Dictionary to store longest count
    longest = {}

    # Initialing dictionary to 0 for all labels
    reset_count(count)
    reset_count(longest)

    # Variable to check last value
    last = -10

    # Iterating complete excel sheet
    for i in range(0, total_count):
        currRow = i+3
        try:
            curr = int(outputSheet.cell(column=11, row=currRow).value)

            # Comparing current and last value
            if(curr==last):
                count[curr]+=1
                longest[curr] = max(longest[curr], count[curr])
                reset_count_except(count, curr)
            else:
                count[curr]=1
                longest[curr] = max(longest[curr], count[curr])
                reset_count_except(count, curr)
        except FileNotFoundError:
            print("File not found!!")
            exit()

        # Updating "last" variable
        last = curr

    # Method to Count longest subsequence frequency
    count_longest_subsequence_freq_func(longest, outputSheet, total_count)

def transition_count_func(row, transition_count, outputSheet):
    # Setting hard coded inputs
    try:
        outputSheet.cell(row=row, column=36).value = "To"
        outputSheet.cell(row=row+1, column=35).value = "Octant #"
        outputSheet.cell(row=row+2, column=34).value = "From"

        for i in range(35, 44):
            for j in range(row+1, row+1+9):
                outputSheet.cell(row=j, column = i).border = black_border

    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()

    # Setting Labels
    for i, label in enumerate(octant_sign):
        try:
            outputSheet.cell(row=row+1, column=i+36).value=label
            outputSheet.cell(row=row+i+2, column=35).value=label
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

    # Setting data
    for i, l1 in enumerate(octant_sign):
        maxi = -1

        for j, l2 in enumerate(octant_sign):
            val = transition_count[str(l1)+str(l2)]
            maxi = max(maxi, val)

        for j, l2 in enumerate(octant_sign):
            try:
                outputSheet.cell(row=row+i+2, column=36+j).value = transition_count[str(l1)+str(l2)]
                if transition_count[str(l1)+str(l2)] == maxi:
                    maxi = -1
                    outputSheet.cell(row=row+i+2, column=36+j).fill = yellow_bg
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

def set_mod_overall_transition_count(outputSheet, mod, total_count):
	# Counting partitions w.r.t. mod
    try:
        totalPartition = total_count//mod
    except ZeroDivisionError:
        print("Mod can't have 0 value")
        exit()

    # Checking mod value range
    if(mod<0):
        raise Exception("Mod value should be in range of 1-30000")

    if(total_count%mod!=0):
        totalPartition +=1

    # Initializing row start for data filling
    rowStart = 16

    # Iterating all partitions 
    for i in range (0,totalPartition):
        # Initializing start and end values
        start = i*mod
        end = min((i+1)*mod-1, total_count-1)

        # Setting start-end values
        try:
            outputSheet.cell(column=35, row=rowStart-1 + 13*i).value = "Mod Transition Count"
            outputSheet.cell(column=35, row=rowStart + 13*i).value = str(start) + "-" + str(end)
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        # Initializing empty dictionary
        transCount = {}
        for a in range (1,5):
            for b in range(1,5):
                transCount[str(a)+str(b)]=0
                transCount[str(a)+str(-b)]=0
                transCount[str(-a)+str(b)]=0
                transCount[str(-a)+str(-b)]=0
                
        # Counting transition for range [start, end)
        for a in range(start, end+1):
            try:
                curr = outputSheet.cell(column=11, row=a+3).value
                next = outputSheet.cell(column=11, row=a+4).value
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

            # Incrementing count for within range value
            if(next!=None):
                transCount[str(curr) + str(next)]+=1

        # Setting transition counts
        transition_count_func(rowStart + 13*i, transCount, outputSheet)

def set_overall_Transition_Count(outputSheet, total_count):
	# Initializing empty dictionary
    count_transition = {}
    for i in range (1,5):
        for j in range(1,5):
            count_transition[str(i)+str(j)]=0
            count_transition[str(i)+str(-j)]=0
            count_transition[str(-i)+str(j)]=0
            count_transition[str(-i)+str(-j)]=0
        
    # Iterating octants values to fill dictionary
    start = 0

    # try and except block for string to int conversion
    try:
        last = int(outputSheet["K3"].value)
    except ValueError:
        print("Sheet input can't be converted to int")
        exit()
    except TypeError:
        print("Sheet doesn't contain integer octant")
        exit()

    while(start<total_count-1):
        # try and except block for string to int conversion
        try:
            curr = int(outputSheet.cell(row= start+4, column=11).value)
            count_transition[str(last) + str(curr)]+=1
            last = curr
        except ValueError:
            print("Sheet input can't be converted to int")
            exit()
        except TypeError:
            print("Sheet doesn't contain integer octant")
            exit()

        start += 1
    
    # Setting transitions counted into sheet
    transition_count_func(2, count_transition, outputSheet)

def set_rank_count(row,countMap, outputSheet):
    # Copying the count list to sort
    sortedCount = []
    count = []
    for label in octant_sign:
        count.append(countMap[label])

    for ct in count:
        sortedCount.append(ct)

    sortedCount.sort(reverse=True)

    rank = []

    for i, el in enumerate(count):
        for j, ell in enumerate(sortedCount):
            if(ell==el):
                rank.append(j+1)
                sortedCount[j] = -1
                break
    rank1Oct = -10

    for j in range(0,8):
        outputSheet.cell(row = row, column=23+j).value = rank[j]
        if(rank[j]==1):
            rank1Oct = octant_sign[j]
            outputSheet.cell(row = row, column=23+j).fill = yellow_bg    

    outputSheet.cell(row=row , column=31).value = rank1Oct
    outputSheet.cell(row=row , column=32).value = octant_name_id_mapping[rank1Oct]

def overall_octant_rank_func(last_row, outputSheet):
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    row =4
    while outputSheet.cell(row=row, column=29).value is not None:
        oct = int(outputSheet.cell(row=row, column=31).value)
        count[oct]+=1
        row+=1

    for i in range(9):
        for j in range(3):
            row = last_row+2+i
            col = 29+j
            outputSheet.cell(row=row, column = col).border = black_border

    outputSheet.cell(column=29, row=last_row+2).value = "Octant ID"
    outputSheet.cell(column=30, row=last_row+2).value = "Octant Name "
    outputSheet.cell(column=31, row=last_row+2).value = "Count of Rank 1 Mod Values"

    for j, oct in enumerate(octant_sign):
        outputSheet.cell(column=29, row=last_row+3+j).value = oct
        outputSheet.cell(column=30, row=last_row+3+j).value = octant_name_id_mapping[oct]
        outputSheet.cell(column=31, row=last_row+3+j).value = count[oct]

def set_mod_count(outputSheet, mod, total_count):
	# Initializing empty dictionary
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    # Variable to store last row
    last_row = -1

    # Iterating loop to set count dictionary
    start = 0
    while(start<total_count):
        count[int(outputSheet.cell(row=start+3, column=11).value)] +=1
        start+=1
        try:
            if(start%mod==0):
                # Setting row data
                try:
                    row = 4 + start//mod
                    last_row = row
                    outputSheet.cell(row=row, column=14).value = str(start-mod) + "-" + str(min(total_count, start-1))

                    for i, label in enumerate(octant_sign):
                        outputSheet.cell(row=row, column=15+i).value = count[label]

                    set_rank_count(row,count, outputSheet)
                except FileNotFoundError:
                    print("Output file not found!!")
                    exit()
                except ValueError:
                    print("Row or column values must be at least 1 ")
                    exit()

                # Reset count values
                count = {-1:0, 1:0,  -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}
        except ZeroDivisionError:
            print("Mod can't have 0 value")
            exit()
    try:
        if(start%mod!=0):
            # Setting row data
            try:
                row = 5 + start//mod
                last_row = row
                outputSheet.cell(row=row, column=14).value = str(start-mod) + "-" + str(min(total_count, start-1))
                for i, label in enumerate(octant_sign):
                    outputSheet.cell(row=row, column=15+i).value = count[label]
                
                set_rank_count(row,count, outputSheet)
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

    except ZeroDivisionError:
        print("Mod can't have 0 value")
        exit()

    if(last_row!=-1):
        overall_octant_rank_func(last_row, outputSheet)

def setOverallCount(total_count, outputSheet):	
	# Initializing count dictionary
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}
    # Incrementing count dictionary data
    try:
        for i in range (3,total_count+3):
            count[int(outputSheet.cell(column=11, row=i).value)] = count[int(outputSheet.cell(column=11, row=i).value)] +1
    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Sheet input can't be converted to int or row/colum should be atleast 1")
        exit()
    except TypeError:
        print("Sheet doesn't contact valid octant value!!")
        exit()

    # Here we are Setting data into sheet
    for i, label in enumerate(octant_sign):
        try:
            outputSheet.cell(row=4, column=i+15).value = count[label]
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

    set_rank_count(4, count, outputSheet)

def set_overall_octant_rank_count(outputSheet, mod, total_count):
    headers = ["Octant ID",1,-1,2,-2,3,-3,+4,-4,"Rank Octant 1", "Rank Octant -1","Rank Octant 2","Rank Octant -2","Rank Octant 3","Rank Octant -3","Rank Octant 4","Rank Octant -4","Rank1 Octant ID","Rank1 Octant Name"]

    totalRows = total_count//mod+1+1 # header + overall
    if total_count%mod!=0:
        totalRows+=1

    for i, header in enumerate(headers):
        for j in range(totalRows):
            outputSheet.cell(row=3+j, column = 14+i).border = black_border

    for i, header in enumerate(headers):
        outputSheet.cell(row=3, column = i+14).value = header

    outputSheet.cell(row=4, column = 13).value = "Mod " + str(mod)

    setOverallCount(total_count, outputSheet)

def get_octant(x,y,z):
    if(x>=0 and y>=0):
        if(z>=0):
            return 1
        else:
            return -1
    
    if(x<0 and y>=0):
        if(z>=0):
            return 2
        else:
            return -2

    if(x<0 and y<0):
        if(z>=0):
            return 3
        else:
            return -3

    if(x>=0 and y<0):
        if(z>=0):
            return 4
        else:
            return -4

def setProcessedDataWithOctant(u_avg, v_avg, w_avg, total_count, inputSheet, outputSheet):
    start = 2
    time = inputSheet.cell(start, 1).value

    # Iterating through out the sheet
    while(time!=None):
        # Calculating processed data
        try:
            u1 = inputSheet.cell(start, 2).value - u_avg
            v1 = inputSheet.cell(start, 3).value - v_avg
            w1 = inputSheet.cell(start, 4).value - w_avg
            
            u1 = round(u1,3)
            v1 = round(v1,3)
            w1 = round(w1,3)

            oct = get_octant(u1, v1, w1)
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        # Setting processed data in the subsequent steps
        try:
            outputSheet.cell(row=start+1, column=8).value = u1
            outputSheet.cell(row=start+1, column=9).value = v1
            outputSheet.cell(row=start+1, column=10).value = w1
            outputSheet.cell(row=start+1, column=11).value = oct
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        start = start+1
        try:
            time = inputSheet.cell(start, 1).value
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()
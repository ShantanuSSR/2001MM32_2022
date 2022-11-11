from datetime import datetime
start_time = datetime.now()

#importing input excel file
import openpyxl
wb = openpyxl.load_workbook(r'octant_input.xlsx')
sheet = wb.active

#finding the number of rows
row_count=sheet.max_row
entire_count=row_count-1


# creating a list to store octants sign
Octant_Sign_List = [1, -1, 2, -2, 3, -3, 4, -4]

#Code
def octant_ranking(mod):
    if mod>30000:
        raise Exception('Mod value needs to be less or equal to 30000.')

    octant_id = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
    variable_row = (30000//mod)+8
    for i, label in enumerate(Octant_Sign_List):
        try:
            sheet.cell(row=1, column=i+22).value = label
            sheet.cell(row=2, column=i+22).value = 'Rank of ' + str(label)
            sheet.cell(row=variable_row+i+1,column=14).value=label
            sheet.cell(row=variable_row+i+1,column=15).value=octant_id[str(label)]
        except FileNotFoundError:
            print("File not found")
            exit()
    try:
        sheet.cell(row=2, column=30).value = 'Rank1 Octant ID'
        sheet.cell(row=2, column=31).value = 'Rank1 Octant Name'
        sheet.cell(row=variable_row, column=14).value = 'Octant ID'
        sheet.cell(row=variable_row, column=15).value = 'Octant Name'
        sheet.cell(row=variable_row, column=16).value = 'Count of Rank1 Mod values'
    except FileNotFoundError:
        print("File not found")
        exit()
    # code for ranking the octants
    for i in range(3,5+(30000//mod)):
        try:
            if i==4:
                continue
            list_tuple_rank=[]
            for j in range(14,22):
                count = int(sheet.cell(row=i,column=j).value)
                column_no = j
                list_tuple_rank.append((column_no,count))
            # sorting in descending order on the basis of count values
            list_tuple_rank.sort(key=lambda x:x[1], reverse=True)
            # storing the rank in the sheet
            rank=1
            for pair in list_tuple_rank:
                sheet.cell(row = i, column = pair[0]+8).value = rank
                rank+=1
            # finding and storing rank1 for each mod value and overall count as well
            for j in range(22,30):
                if int(sheet.cell(row = i, column = j).value) == 1:
                    octant_sign_rank1 = sheet.cell(row = 1, column = j).value
                    sheet.cell(row = i, column = 30).value = octant_sign_rank1
                    sheet.cell(row = i, column = 31).value = octant_id[str(octant_sign_rank1)]
                    break
        except FileNotFoundError:
            print("File not found")
            exit()
    # calculate the number of rank1 for each octant_sign per mod
    begin = variable_row+1
    for j in range(22,30):
        try:
            rank1_count = 0
            for i in range(5,5+(30000//mod)):
                if int(sheet.cell(row = i,column = j).value)==1:
                    rank1_count+=1
            sheet.cell(row = begin,column = 16).value = rank1_count
            begin+=1
        except FileNotFoundError:
            print ("File missing")
            exit()

def count_in_range(mod):
    if mod>30000:
        raise Exception('mod value should be less than or equal to 30000')
    begin=2
    add=5
    while begin<entire_count:
        # loop to initialize the cell value as 0
        try:
            for j in range(14,22):
                sheet.cell(row=add,column=j).value=0

            # loop to add the count of octant in appropriate cell
            for i in range(begin,min(row_count+1,mod+begin)):
                region = sheet.cell(row=i,column=11).value
                if region==1:
                    sheet.cell(row=add,column=14).value += 1
                elif region==-1:
                    sheet.cell(row=add,column=15).value+=1
                elif region==2:
                    sheet.cell(row=add,column=16).value+=1
                elif region==-2:
                    sheet.cell(row=add,column=17).value+=1
                elif region==3:
                    sheet.cell(row=add,column=18).value+=1
                elif region==-3:
                    sheet.cell(row=add,column=19).value+=1
                elif region==4:
                    sheet.cell(row=add,column=20).value+=1
                else:
                    sheet.cell(row=add,column=21).value+=1
            # adding range value in appropriate cell
            x1=str(begin-2)
            y1=str(min(entire_count-1,mod+begin-3))
            sheet.cell(row=add,column=13).value=x1+'-'+y1

            # changed the value of for by increasing it by mod
            begin+=mod
            #increasing adding row by 1
            add+=1
        except FileNotFoundError:
            print('File Missing!')
            exit()

def check_octant_sign(u, v, w):
    if u > 0:
        if v > 0:
            if w > 0:
                # (u,v) is +ve and w is +ve so +1
                return 1
            else:
                # (u,v) is +ve  but w is -ve so -1
                return -1
        else:
            if w > 0:
                #  u>0 and v<0 and w is +ve so +4
                return 4
            else:
                #u>0 and v<0 but w is -ve so -4
                return -4
    else:
        if v > 0:
            if w > 0:
                #u<0 and v>0 and w is +ve so +2
                return 2
            else:
                #u<0 and v>0 but w is -ve so -2
                return -2
        else:
            if w > 0:
                #u<0 and v<0 and w is +ve so +3
                return 3
            else:
                #u<0 and v<0 and w is -ve so -3
                return -3

def avg_calc():
    val_U=0
    val_V=0
    val_W=0
    for i in range(2, row_count + 1):
        try:
            val_U += sheet.cell(row=i,column=2).value
            val_V += sheet.cell(row=i,column=3).value
            val_W += sheet.cell(row=i,column=4).value
        except FileNotFoundError:
            print('File not found!')
    sheet['E1']='u_avg'
    sheet['F1']='v_avg'
    sheet['G1']='w_avg'
    # average of u, v, w
    try:
        u_avg=val_U/entire_count
        v_avg=val_V/entire_count
        w_avg=val_W/entire_count
    except(ZeroDivisionError):
        print("No input data found!!\nDivision by zero is not allowed!")
        exit()
    try:
        #saving average of U in the sheet
        sheet['E2']=u_avg
        #saving average of V in the sheet
        sheet['F2']=v_avg
        #saving average of W in the sheet
        sheet['G2']=w_avg
    except FileNotFoundError:
        print("File not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()

    sheet['H1']='U-u_avg'
    sheet['I1']='V-v_avg'
    sheet['J1']='W-w_avg'
    for i in range(2, row_count + 1):
        #calculating and saving U-u_avg in the sheet
        sub_u_avg = sheet.cell(row=i,column=2).value-u_avg
        sheet.cell(row=i,column=8).value=sub_u_avg

        #calculating and saving V-v_avg in the sheet
        sub_v_avg = sheet.cell(row=i,column=3).value-v_avg
        sheet.cell(row=i,column=9).value=sub_v_avg

        #calculating and saving W-w_avg in the sheet
        sub_w_avg = sheet.cell(row=i,column=4).value-w_avg
        sheet.cell(row=i,column=10).value=sub_w_avg

def octant_identification(mod):
    if mod>30000:
        raise Exception('mod value should be less than or equal to 30000')
    #function to calculate and save average value of U, V, W
    avg_calc()

    sheet['K1']='Octant'
    #initializing count values of each octant sign as 0
    for j in range(14,22):
        sheet.cell(row=3,column=j).value=0
    #saving the sign of the octant
    for i in range(2,row_count+1):
        try:
            sub_u_avg=sheet.cell(row=i,column=8).value
            sub_v_avg=sheet.cell(row=i,column=9).value
            sub_w_avg=sheet.cell(row=i,column=10).value
            try:
                octant_sign=check_octant_sign(sub_u_avg,sub_v_avg,sub_w_avg)
            except NameError:
                print('Either the function is not defined or is not named correctly')
                exit()
            if octant_sign==1:
                sheet.cell(row=3,column=14).value+=1
            elif octant_sign==-1:
                sheet.cell(row=3,column=15).value+=1
            elif octant_sign==2:
                sheet.cell(row=3,column=16).value+=1
            elif octant_sign==-2:
                sheet.cell(row=3,column=17).value+=1
            elif octant_sign==3:
                sheet.cell(row=3,column=18).value+=1
            elif octant_sign==-3:
                sheet.cell(row=3,column=19).value+=1
            elif octant_sign==4:
                sheet.cell(row=3,column=20).value+=1
            elif octant_sign==-4:
                sheet.cell(row=3,column=21).value+=1
            sheet.cell(row=i,column=11).value=octant_sign
        except FileNotFoundError:
            print('File not found!')
            exit()
    sheet['L4']='user input'
    sheet['M2']='Octant ID'
    sheet['M3']='Overall Count'
    sheet['M4']= 'Mod '+ str(mod)
    # creating headng of the columns with assing value as 1, -1, 2, -2, 3, -3, 4, -4 
    for j, label in enumerate(Octant_Sign_List):
        sheet.cell(row=2, column=j+14).value=label
        
    # calling count_in_range function to find the octant sign
    count_in_range(mod)